"""
CSV Pipeline Module for SID Manager
Handles: CSV cleaning (Indeed/LinkedIn), master list management,
30-day phone deduplication, and row division for HeyMarket/GoHighLevel.
Ported from the standalone CSV Tools desktop app.
"""

import re
import io
import csv
import json
import os
import random
import unicodedata
import datetime
import zipfile
from pathlib import Path

import pandas as pd

# =============================================================================
# CONFIGURATION
# =============================================================================

PIPELINE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pipeline_data")
MASTER_LISTS_DIR = os.path.join(PIPELINE_DIR, "master_lists")
CLEANED_DIR = os.path.join(PIPELINE_DIR, "cleaned")
DIVIDED_DIR = os.path.join(PIPELINE_DIR, "divided")
CONFIG_PATH = os.path.join(PIPELINE_DIR, "pipeline_config.json")

# Ensure directories exist
for d in [PIPELINE_DIR, MASTER_LISTS_DIR, CLEANED_DIR, DIVIDED_DIR]:
    os.makedirs(d, exist_ok=True)


# =============================================================================
# SHARED UTILITIES (ported from CSV Tools app.py)
# =============================================================================

ENCODINGS_TO_TRY = [
    "utf-8-sig", "utf-8",
    "utf-16", "utf-16le", "utf-16be",
    "utf-32", "utf-32le", "utf-32be",
    "cp1252", "latin-1", "iso-8859-1", "mac_roman",
]
DELIMITERS_TO_TRY = [",", ";", "\t", "|"]

NON_US_NANP_AREA_CODES = {
    "204","226","236","249","250","263","289","306","343","354","365","367","368",
    "387","403","416","418","431","437","438","450","468","474","506","514","519",
    "548","579","581","584","587","604","613","639","647","672","683","705","709",
    "742","753","778","780","782","807","819","825","867","873","879","902","905",
    "242","246","264","268","284","340","345","441","473","649","664","670","671",
    "684","721","758","767","784","787","809","829","849","868","869","876","939",
}

FANCY_PUNCT_MAP = {
    "\u2018": "'", "\u2019": "'", "\u201B": "'", "\u2032": "'", "\u00B4": "'",
    "\u201C": '"', "\u201D": '"',
    "\u2010": "-", "\u2011": "-", "\u2012": "-", "\u2013": "-",
    "\u00A0": " ",
}

# Column name candidates for auto-detection
FIRST_CANDIDATES = ["first name", "first", "fname", "given", "given name"]
LAST_CANDIDATES = ["last name", "last", "lname", "surname", "family name"]
FULL_CANDIDATES = ["name", "full name", "candidate name", "applicant name"]
PHONE_CANDIDATES = ["phone", "phone number", "mobile", "mobile phone", "cell", "cell phone", "telephone", "contact number"]
POS_CANDIDATES = ["position", "job", "job title", "role", "title"]


def nfc(s):
    if s is None:
        return ""
    return unicodedata.normalize("NFC", str(s).replace("\u00A0", " ")).strip()


def norm_text(s, strip_accents=False):
    if s is None:
        return ""
    s = str(s)
    for k, v in FANCY_PUNCT_MAP.items():
        s = s.replace(k, v)
    s = unicodedata.normalize("NFKC", s)
    if strip_accents:
        s = "".join(
            ch for ch in unicodedata.normalize("NFKD", s)
            if not unicodedata.combining(ch)
        )
    return re.sub(r"\s+", " ", s).strip()


def smart_title(name):
    name = nfc(name)
    if not name:
        return ""
    name = re.sub(r"\s+", " ", name)
    parts = name.split(" ")
    output = []
    for p in parts:
        hyphen_parts = p.split("-")
        hyphen_out = []
        for seg in hyphen_parts:
            if not seg:
                continue
            apos = "'" if "'" in seg else ("'" if "\u2019" in seg else None)
            if apos:
                sub = seg.split(apos)
                rebuilt = []
                for i, s in enumerate(sub):
                    if s.lower().startswith("mc") and len(s) > 2:
                        rebuilt.append("Mc" + s[2].upper() + s[3:].lower())
                    elif i == 0 and s.lower() in {"o", "de", "da", "van", "von"}:
                        rebuilt.append(s.lower())
                    else:
                        rebuilt.append(s[:1].upper() + s[1:].lower() if s else s)
                seg = apos.join(rebuilt)
            else:
                if seg.lower().startswith("mc") and len(seg) > 2:
                    seg = "Mc" + seg[2].upper() + seg[3:].lower()
                else:
                    seg = seg[:1].upper() + seg[1:].lower() if seg else seg
            hyphen_out.append(seg)
        output.append("-".join(hyphen_out))
    return " ".join(output)


def split_full_name(name):
    name = nfc(name)
    if not name:
        return "", ""
    if "," in name:
        last, rest = name.split(",", 1)
        first = rest.strip().split(" ")[0] if rest.strip() else ""
        return first, last.strip()
    parts = name.split()
    if len(parts) == 1:
        return parts[0], "Unknown"
    return parts[0], " ".join(parts[1:])


def normalize_col(col):
    return re.sub(r"[^a-z0-9]+", " ", col.lower()).strip()


def find_col(df, candidates):
    for c in df.columns:
        norm = normalize_col(str(c))
        for cand in candidates:
            if cand == norm or cand in norm:
                return c
    return None


def find_col_by_alias(cols, aliases):
    for c in cols:
        cl = c.lower().replace(" ", "")
        for a in aliases:
            if cl == a.replace(" ", ""):
                return c
    return None


def infer_position_from_filename(name):
    base = Path(name).stem
    base = re.sub(r"_candidates$", "", base, flags=re.I)
    return base.replace("_", " ").strip()


def norm_phone_indeed(raw):
    raw = nfc(raw)
    if not raw:
        return None, "no_phone"
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"+1 {digits}", ""
    if len(digits) >= 12 or (len(digits) == 11 and not digits.startswith("1")):
        return None, "international"
    return None, "invalid"


def normalize_phone_linkedin(raw):
    digits = re.sub(r"\D+", "", str(raw))
    if len(digits) == 10:
        return None if digits[:3] in NON_US_NANP_AREA_CODES else "+1 " + digits
    if len(digits) == 11 and digits.startswith("1"):
        ten = digits[-10:]
        return None if ten[:3] in NON_US_NANP_AREA_CODES else "+1 " + ten
    return None


def read_csv_robust(file_bytes):
    """Parse CSV bytes with multiple encoding/delimiter attempts."""
    for enc in ENCODINGS_TO_TRY:
        try:
            text = file_bytes.decode(enc)
        except Exception:
            continue
        sniffed = None
        try:
            dialect = csv.Sniffer().sniff(text[:4000], delimiters="".join(DELIMITERS_TO_TRY))
            sniffed = dialect.delimiter
        except Exception:
            sniffed = None
        delims = ([sniffed] if sniffed else []) + DELIMITERS_TO_TRY
        for d in delims:
            try:
                df = pd.read_csv(
                    io.StringIO(text), engine="python", dtype=str,
                    keep_default_na=False, sep=d, on_bad_lines="skip",
                )
                if df.shape[1] > 1:
                    return df
            except Exception:
                continue
    text = file_bytes.decode("utf-8", errors="replace")
    for d in DELIMITERS_TO_TRY:
        try:
            df = pd.read_csv(
                io.StringIO(text), engine="python", dtype=str,
                keep_default_na=False, sep=d, on_bad_lines="skip",
            )
            if df.shape[1] > 1:
                return df
        except Exception:
            continue
    raise RuntimeError("Unable to parse CSV")


def safe_filename(name):
    return "".join(c if c.isalnum() or c in ["_", "-"] else "_" for c in name)


# =============================================================================
# INDEED CLEANER
# =============================================================================

def clean_indeed(csv_bytes_list, company="", list_val="", date_val="", stage="Initial"):
    """
    Clean one or more Indeed CSV byte buffers.
    Returns: (cleaned_df, stats_dict)
    """
    if not date_val:
        today = datetime.date.today()
        date_val = f"{today.month}/{today.day}/{today.year}"

    frames = []
    per_file = {}

    for fname, fbytes in csv_bytes_list:
        try:
            if fname.lower().endswith((".xlsx", ".xls")):
                xls = pd.ExcelFile(io.BytesIO(fbytes))
                for sheet in xls.sheet_names:
                    temp = xls.parse(sheet, dtype=str, keep_default_na=False)
                    if not temp.empty:
                        per_file[fname] = len(temp)
                        frames.append(_clean_df_indeed(temp, fname, company, list_val, date_val, stage))
                        break
                else:
                    per_file[fname] = 0
            else:
                df = read_csv_robust(fbytes)
                per_file[fname] = len(df)
                frames.append(_clean_df_indeed(df, fname, company, list_val, date_val, stage))
        except Exception as e:
            per_file[fname] = f"error: {str(e)}"

    if not frames:
        return None, {"error": "No data loaded", "per_file": per_file}

    combined = pd.concat(frames, ignore_index=True)
    total_before = len(combined)

    dropped_international = int((combined["_drop"] == "international").sum())
    dropped_no_phone = int((combined["_drop"] == "no_phone").sum())
    dropped_invalid = int((combined["_drop"] == "invalid").sum())

    kept = combined[combined["_drop"] == ""].copy()
    before_dedupe = len(kept)
    kept = kept.drop_duplicates(subset=["Phone Number"], keep="first").copy()
    dropped_dupes = before_dedupe - len(kept)

    final = kept[["First Name", "Last Name", "Phone Number", "List", "Position", "Company", "Date", "Stage"]].copy()

    stats = {
        "per_file": per_file,
        "total_rows": total_before,
        "dropped_dupes": dropped_dupes,
        "dropped_international": dropped_international,
        "dropped_no_phone": dropped_no_phone,
        "dropped_invalid": dropped_invalid,
        "final_rows": len(final),
    }

    return final, stats


def _clean_df_indeed(df, source, company, list_val, date_val, stage):
    df = df.copy()
    df.columns = [nfc(c) for c in df.columns]

    c_first = find_col(df, FIRST_CANDIDATES)
    c_last = find_col(df, LAST_CANDIDATES)
    c_full = find_col(df, FULL_CANDIDATES)
    c_phone = find_col(df, PHONE_CANDIDATES)
    c_pos = find_col(df, POS_CANDIDATES)

    first, last = [], []
    for i in range(len(df)):
        if c_first is not None and c_last is not None:
            f = df.iloc[i][c_first]
            l = df.iloc[i][c_last]
            if (" " in str(f)) or ("," in str(f)):
                f, l = split_full_name(f)
            elif (" " in str(l)) or ("," in str(l)):
                f, l = split_full_name(l)
        elif c_full is not None:
            f, l = split_full_name(df.iloc[i][c_full])
        else:
            f, l = "", ""
        first.append(smart_title(f))
        last.append(smart_title(l))

    phones, drops = [], []
    phone_series = df[c_phone].astype(str) if c_phone is not None else pd.Series([""] * len(df))
    for v in phone_series.tolist():
        p, r = norm_phone_indeed(v)
        phones.append(p or "")
        drops.append(r or "")

    if c_pos is not None:
        pos = df[c_pos].astype(str)
    else:
        pos = pd.Series([infer_position_from_filename(source)] * len(df))
    pos = pos.astype(str).str.replace(r"\s+- Work From Home\s*$", "", regex=True)

    return pd.DataFrame({
        "First Name": first,
        "Last Name": last,
        "Phone Number": phones,
        "List": list_val,
        "Position": pos,
        "Company": company,
        "Date": date_val,
        "Stage": stage,
        "_drop": drops,
        "_src": source,
    })


# =============================================================================
# LINKEDIN CLEANER
# =============================================================================

def clean_linkedin(csv_bytes_list, company="FBSPL", list_val="", date_val="", stage="Start", strip_accents=False):
    """
    Clean one or more LinkedIn CSV/XLSX byte buffers.
    Returns: (cleaned_df, stats_dict)
    """
    if not date_val:
        today = datetime.date.today()
        date_val = f"{today.month}/{today.day}/{today.year}"

    all_cleaned = []
    stat_rows = []

    for fname, fbytes in csv_bytes_list:
        try:
            if fname.lower().endswith((".xlsx", ".xls")):
                xls = pd.ExcelFile(io.BytesIO(fbytes))
                sheet = _pick_sheet(xls)
                df_raw = pd.read_excel(io.BytesIO(fbytes), sheet_name=sheet, dtype=str, keep_default_na=False)
            else:
                df_raw = read_csv_robust(fbytes)

            cleaned, stats = _clean_df_linkedin(df_raw, company, list_val, stage, date_val, strip_accents)
            stats["file"] = fname
            stat_rows.append(stats)
            all_cleaned.append(cleaned)
        except Exception as e:
            stat_rows.append({"file": fname, "rows_in": 0, "intl_removed": 0, "rows_kept": 0, "error": str(e)})

    if not all_cleaned:
        return None, {"error": "No data loaded", "file_stats": stat_rows}

    combined = pd.concat(all_cleaned, ignore_index=True)
    before_dedupe = len(combined)
    combined = combined.drop_duplicates(subset=["Phone Number"], keep="first")
    dupes_removed = before_dedupe - len(combined)

    total_stats = {
        "file_stats": stat_rows,
        "total_rows_in": sum(s.get("rows_in", 0) for s in stat_rows),
        "total_intl_removed": sum(s.get("intl_removed", 0) for s in stat_rows),
        "rows_before_dedupe": before_dedupe,
        "dupes_removed": dupes_removed,
        "final_rows": len(combined),
    }

    return combined, total_stats


def _pick_sheet(xls):
    for s in xls.sheet_names:
        if "job" in s.lower():
            return s
    return xls.sheet_names[0]


def _clean_df_linkedin(df, company, list_val, stage, date_str, strip_accents):
    original_rows = len(df)
    df = df.copy()
    df.columns = [norm_text(c, strip_accents) for c in df.columns]
    for c in df.columns:
        df[c] = df[c].map(lambda x: norm_text(x, strip_accents))

    phone_col = next((c for c in df.columns if any(x in c.lower() for x in ["phone", "mobile", "contact"])), None)
    phones = df[phone_col].map(normalize_phone_linkedin) if phone_col else pd.Series([None] * len(df))

    valid_mask = phones.notna()
    intl_removed = original_rows - valid_mask.sum()

    df = df[valid_mask].reset_index(drop=True)
    phones = phones[valid_mask].reset_index(drop=True)

    first_col = find_col_by_alias(df.columns, ["firstname", "first", "fname", "given"])
    last_col = find_col_by_alias(df.columns, ["lastname", "last", "lname", "surname"])
    full_col = find_col_by_alias(df.columns, ["name", "fullname", "candidatename", "applicantname"])

    if first_col and last_col:
        first = df[first_col]
        last = df[last_col]
    else:
        src = df[full_col] if full_col else pd.Series([""] * len(df))
        split = src.map(lambda x: x.split(" ", 1))
        first = split.map(lambda x: x[0] if x else "")
        last = split.map(lambda x: x[1] if len(x) > 1 else "")

    def _proper(x):
        x = norm_text(x, strip_accents)
        if not x:
            return "Unknown"
        parts = []
        for p in x.split():
            if p.lower().startswith("mc") and len(p) > 2:
                parts.append("Mc" + p[2:].capitalize())
            elif "'" in p:
                segs = p.split("'")
                parts.append("'".join(s.capitalize() for s in segs))
            else:
                parts.append(p.capitalize())
        return " ".join(parts)

    first = first.map(_proper)
    last = last.map(_proper)

    pos_col = next((c for c in df.columns if "job" in c.lower() and "title" in c.lower()), None)
    position = df[pos_col] if pos_col else "Unknown"

    cleaned = pd.DataFrame({
        "First Name": first,
        "Last Name": last,
        "Phone Number": phones,
        "List": list_val,
        "Position": position,
        "Company": company,
        "Date": date_str,
        "Stage": stage,
    })

    stats = {"rows_in": original_rows, "intl_removed": int(intl_removed), "rows_kept": len(cleaned)}
    return cleaned, stats


# =============================================================================
# MASTER LIST MANAGEMENT — Excel Workbook with Rolling Tab + Monthly Archives
# =============================================================================
#
# Structure per category:
#   SFG → Indeed_SFG_MasterList.xlsx
#     "30 Days" tab  — rolling dedup window (new data appended here, deduped against this)
#     "Mar 26" tab   — archived data from March 2026
#     "Feb 26" tab   — archived data from February 2026
#     ...
#
#   MS → Slice_Indeed_MasterList.xlsx
#     "60 Days" tab  — rolling 60-day dedup window
#     "Mar 26" tab, "Feb 26" tab, ...
#
# Columns: First Name, Last Name, Phone Number, List, Position, Company, Date, Stage
# Dates are stored as Excel datetime objects.

MASTER_LIST_COLS = ["First Name", "Last Name", "Phone Number", "List", "Position", "Company", "Date", "Stage"]

# Category config: maps category key to file path, rolling tab name, dedup window
# Indeed and LinkedIn each have their own master list per category
_ONEDRIVE = os.path.expanduser("~/Library/CloudStorage/OneDrive-JobosaurusVentureHoldingCoInc")

CATEGORY_CONFIGS = {
    # Indeed master lists
    "SFG": {
        "file": os.path.join(_ONEDRIVE, "Indeed_SFG_MasterList.xlsx"),
        "rolling_tab": "30 Days",
        "dedup_days": 30,
        "archive_mode": "move",
        "source": "indeed",
    },
    "MS": {
        "file": os.path.join(_ONEDRIVE, "Slice_Indeed_MasterList.xlsx"),
        "rolling_tab": "60 Days",
        "dedup_days": 60,
        "archive_mode": "copy",
        "source": "indeed",
    },
    # LinkedIn master lists
    "SFG_LinkedIn": {
        "file": os.path.join(_ONEDRIVE, "LinkedIn_SFG_MasterList.xlsx"),
        "rolling_tab": "30 Days",
        "dedup_days": 30,
        "archive_mode": "move",
        "source": "linkedin",
    },
    "MS_LinkedIn": {
        "file": os.path.join(_ONEDRIVE, "Slice_LinkedIn_MasterList.xlsx"),
        "rolling_tab": "60 Days",
        "dedup_days": 60,
        "archive_mode": "copy",
        "source": "linkedin",
    },
}

# Base directory for SID data (postings, assignments, etc.)
SID_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Daily pipeline results log
PIPELINE_LOG_PATH = os.path.join(PIPELINE_DIR, "pipeline_log.json")


def _load_pipeline_log():
    if os.path.exists(PIPELINE_LOG_PATH):
        try:
            with open(PIPELINE_LOG_PATH, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {"runs": []}


def _save_pipeline_log(log):
    with open(PIPELINE_LOG_PATH, "w") as f:
        json.dump(log, f, indent=2, default=str)


def log_pipeline_run(entry):
    log = _load_pipeline_log()
    log["runs"].append(entry)
    if len(log["runs"]) > 200:
        log["runs"] = log["runs"][-200:]
    _save_pipeline_log(log)


def get_pipeline_log(date_filter=None):
    log = _load_pipeline_log()
    if date_filter:
        log["runs"] = [r for r in log["runs"] if r.get("date", "").startswith(date_filter)]
    return log


def _load_vendor_postings(vendor):
    """Load posting templates for a vendor, returning {category: [title, ...]}."""
    fpath = os.path.join(SID_BASE_DIR, f"postings_{vendor}.json")
    if not os.path.exists(fpath):
        return {}
    try:
        with open(fpath, "r") as f:
            data = json.load(f)
        cat_titles = {}
        for cat, posts in data.get("categories", {}).items():
            titles = [p.get("title", "") for p in posts if p.get("title")]
            if titles:
                cat_titles[cat] = titles
        return cat_titles
    except Exception:
        return {}


def _load_vendor_assignments(vendor, lookback_days=14):
    """Load recent assignments for a vendor, returning {title: category}."""
    assign_dir = os.path.join(SID_BASE_DIR, "assignments", vendor.lower())
    if not os.path.exists(assign_dir):
        return {}
    cutoff = datetime.date.today() - datetime.timedelta(days=lookback_days)
    title_to_cat = {}
    try:
        for fname in sorted(os.listdir(assign_dir), reverse=True):
            if not fname.endswith(".json"):
                continue
            date_str = fname.replace(".json", "")
            try:
                fdate = datetime.date.fromisoformat(date_str)
            except ValueError:
                continue
            if fdate < cutoff:
                continue
            with open(os.path.join(assign_dir, fname), "r") as f:
                data = json.load(f)
            for a in data.get("assignments", []):
                title = a.get("title", "")
                cat = a.get("category", "")
                # Skip generic "PUSH" category — use the category from the assignment ID/uid instead
                if cat == "PUSH":
                    uid = a.get("uid", "") or a.get("id", "")
                    # UIDs look like "FBSPL-SFG-0-1774713897788" — category is 2nd segment
                    parts = uid.split("-")
                    if len(parts) >= 2:
                        cat = parts[1] if parts[0] == vendor else (parts[0] if parts[0] in CATEGORY_CONFIGS else "")
                if title and cat and cat in CATEGORY_CONFIGS:
                    title_to_cat[title] = cat
    except Exception:
        pass
    return title_to_cat


def _fuzzy_title_match(csv_title, candidates, threshold=0.5):
    """Jaccard token similarity with substring bonus. Returns (best_match, score)."""
    if not csv_title or not candidates:
        return None, 0

    def tokenize(s):
        return set(re.sub(r'[^a-z0-9\s]', '', s.lower()).split())

    csv_tokens = tokenize(csv_title)
    if not csv_tokens:
        return None, 0

    best_match = None
    best_score = 0

    for candidate in candidates:
        cand_tokens = tokenize(candidate)
        if not cand_tokens:
            continue
        intersection = csv_tokens & cand_tokens
        union = csv_tokens | cand_tokens
        score = len(intersection) / len(union) if union else 0

        # Bonus: if one is a substring of the other
        if csv_title.lower() in candidate.lower() or candidate.lower() in csv_title.lower():
            score = max(score, 0.85)

        if score > best_score:
            best_score = score
            best_match = candidate

    if best_score >= threshold:
        return best_match, best_score
    return None, 0


def _build_title_to_category_map(vendor):
    """
    Build a comprehensive title -> category mapping for a vendor by combining:
    1. Posting templates (postings_{vendor}.json) — canonical categories
    2. Recent assignments (assignments/{vendor}/*.json) — operational categories
    Returns {title: category} and a flat list of (title, category) pairs for fuzzy matching.
    """
    title_to_cat = {}

    # 1. Load posting templates — these are the canonical source
    posting_cats = _load_vendor_postings(vendor)
    for cat, titles in posting_cats.items():
        if cat in CATEGORY_CONFIGS:
            for t in titles:
                title_to_cat[t] = cat

    # 2. Load recent assignments — fill in any missing titles
    assign_cats = _load_vendor_assignments(vendor)
    for title, cat in assign_cats.items():
        if title not in title_to_cat:  # Don't override posting template categories
            title_to_cat[title] = cat

    return title_to_cat


def categorize_cleaned_data(vendor, cleaned_df, source="indeed"):
    """
    Split a cleaned DataFrame into per-category DataFrames based on matching
    each row's Position against the vendor's postings/assignments.

    source: "indeed" or "linkedin" — determines which master list configs to use.
    For LinkedIn data, base categories (SFG, MS) get mapped to SFG_LinkedIn, MS_LinkedIn.

    Returns: dict of {category: DataFrame}, plus a list of unmatched rows.
    """
    title_to_cat = _build_title_to_category_map(vendor)

    if not title_to_cat:
        # No postings/assignments configured for this vendor — can't categorize
        return {}, cleaned_df

    all_titles = list(title_to_cat.keys())
    pos_col = None
    for col in cleaned_df.columns:
        if col.lower().strip() in ["position", "job title", "title", "job", "role"]:
            pos_col = col
            break

    if pos_col is None:
        # No position column — can't match, return all as uncategorized
        return {}, cleaned_df

    # Build a category column by matching each row's position to known titles
    categories = []
    for pos in cleaned_df[pos_col]:
        pos = str(pos).strip() if pd.notna(pos) else ""

        # Try exact match first
        if pos in title_to_cat:
            categories.append(title_to_cat[pos])
            continue

        # Try fuzzy match
        match, score = _fuzzy_title_match(pos, all_titles, threshold=0.5)
        if match:
            categories.append(title_to_cat[match])
        else:
            categories.append(None)

    cleaned_df = cleaned_df.copy()
    cleaned_df["_matched_category"] = categories

    # For LinkedIn source, map base categories to LinkedIn-specific configs
    # e.g. "SFG" -> "SFG_LinkedIn", "MS" -> "MS_LinkedIn"
    suffix = "_LinkedIn" if source == "linkedin" else ""

    # Split by category
    result = {}
    for base_cat in ["SFG", "MS"]:  # The base categories from postings
        target_cat = base_cat + suffix
        if target_cat not in CATEGORY_CONFIGS:
            continue
        cat_df = cleaned_df[cleaned_df["_matched_category"] == base_cat].drop(columns=["_matched_category"])
        if not cat_df.empty:
            result[target_cat] = cat_df.reset_index(drop=True)

    unmatched = cleaned_df[cleaned_df["_matched_category"].isna()].drop(columns=["_matched_category"])
    return result, unmatched.reset_index(drop=True)


def _detect_csv_source(csv_list):
    """
    Detect whether CSVs are from Indeed or LinkedIn based on column headers.
    LinkedIn CSVs typically have columns like 'Company / Organization' or lack 'Position'.
    Indeed CSVs have 'Candidate Name' or 'First Name' etc.
    Returns: "indeed" or "linkedin"
    """
    for fname, fbytes in csv_list:
        try:
            text = fbytes.decode("utf-8-sig", errors="replace")
            header_line = text.split("\n")[0].lower()
            # LinkedIn indicators
            if "company / organization" in header_line or "linkedin" in fname.lower():
                return "linkedin"
            if "member since" in header_line or "profile url" in header_line:
                return "linkedin"
        except Exception:
            continue
    return "indeed"


def auto_process_vendor_upload(vendor, zip_path, upload_type="resumes"):
    """
    Auto-process a vendor's uploaded ZIP.
    Called by the server right after a vendor uploads a 'resumes' ZIP.
    Flow: Extract CSVs -> Detect source (Indeed/LinkedIn) -> Clean
          -> Match titles to postings/categories -> Split by category
          -> Add each to correct master list -> Dedup -> Export
    """
    today = datetime.date.today()
    date_str = f"{today.month}/{today.day}/{today.year}"

    results = {
        "vendor": vendor,
        "date": today.isoformat(),
        "timestamp": datetime.datetime.now().isoformat(),
        "stages": [],
        "categories_processed": {},
    }

    try:
        with open(zip_path, "rb") as f:
            zip_bytes = f.read()

        csv_list = extract_csvs_from_zip(zip_bytes)
        results["stages"].append({
            "name": "extract",
            "files_found": len(csv_list),
            "filenames": [f[0] for f in csv_list],
        })

        if not csv_list:
            results["error"] = "No CSV files found in ZIP"
            log_pipeline_run(results)
            return results

        # Detect source type (Indeed vs LinkedIn) from CSV headers
        source = _detect_csv_source(csv_list)
        results["source"] = source

        # Clean using the appropriate cleaner
        if source == "linkedin":
            cleaned_df, clean_stats = clean_linkedin(
                csv_list, company=vendor, list_val="", date_val=date_str, stage="Initial"
            )
        else:
            cleaned_df, clean_stats = clean_indeed(
                csv_list, company=vendor, list_val="", date_val=date_str, stage="Initial"
            )
        results["stages"].append({"name": "clean", "type": source, **clean_stats})
        results["cleaned_rows"] = len(cleaned_df) if cleaned_df is not None else 0

        if cleaned_df is None or cleaned_df.empty:
            results["error"] = "Cleaning produced no results"
            log_pipeline_run(results)
            return results

        # Match CSV positions to postings/assignments to determine categories
        # source determines which master list configs get used (SFG vs SFG_LinkedIn, etc.)
        cat_dfs, unmatched_df = categorize_cleaned_data(vendor, cleaned_df, source=source)

        results["stages"].append({
            "name": "categorize",
            "method": "posting_title_match",
            "categories_found": {cat: len(df) for cat, df in cat_dfs.items()},
            "unmatched_rows": len(unmatched_df),
        })

        if not cat_dfs:
            results["error"] = f"No rows matched any category postings for {vendor}. " \
                              f"Make sure this vendor has postings organized by category (SFG, MS) in the Postings page. " \
                              f"Unmatched rows: {len(unmatched_df)}"
            results["ok"] = False
            log_pipeline_run(results)
            return results

        # Process each category separately
        total_export_rows = 0
        for category, cat_df in cat_dfs.items():
            cat_result = {"category": category, "rows": len(cat_df)}

            try:
                add_stats = add_to_master_list(category, cat_df)
                cat_result["master_add"] = add_stats

                today_kept, removed, dedup_stats = dedup_master_list(category)
                cat_result["dedup"] = dedup_stats

                if not today_kept.empty:
                    export_dir = os.path.join(CLEANED_DIR, today.isoformat())
                    os.makedirs(export_dir, exist_ok=True)
                    export_path = os.path.join(export_dir, f"{safe_filename(vendor)}_{safe_filename(category)}_export.csv")

                    export_df = today_kept.copy()
                    if "Date" in export_df.columns:
                        export_df["Date"] = export_df["Date"].apply(lambda x: str(x) if x is not None else "")
                    export_df.to_csv(export_path, index=False)

                    cat_result["export_path"] = export_path
                    cat_result["export_rows"] = len(today_kept)
                    total_export_rows += len(today_kept)
                else:
                    cat_result["export_rows"] = 0

                cat_result["ok"] = True

            except Exception as e:
                import traceback as tb
                cat_result["error"] = str(e)
                cat_result["traceback"] = tb.format_exc()

            results["categories_processed"][category] = cat_result

        results["export_rows"] = total_export_rows
        results["category"] = ", ".join(cat_dfs.keys())  # summary for backward compat
        results["ok"] = True

        if len(unmatched_df) > 0:
            results["unmatched_rows"] = len(unmatched_df)
            # Save unmatched for review
            export_dir = os.path.join(CLEANED_DIR, today.isoformat())
            os.makedirs(export_dir, exist_ok=True)
            unmatched_path = os.path.join(export_dir, f"{safe_filename(vendor)}_UNMATCHED.csv")
            unmatched_df.to_csv(unmatched_path, index=False)
            results["unmatched_export_path"] = unmatched_path

    except Exception as e:
        import traceback as tb
        results["error"] = str(e)
        results["traceback"] = tb.format_exc()

    log_pipeline_run(results)
    return results


def get_today_exports():
    """Get all exportable CSVs from today's pipeline runs."""
    today = datetime.date.today().isoformat()
    log = get_pipeline_log(date_filter=today)

    exports = []
    for run in log.get("runs", []):
        if run.get("ok") and run.get("export_path"):
            path = run["export_path"]
            rows = run.get("export_rows", 0)
            if os.path.exists(path):
                try:
                    df = pd.read_csv(path, dtype=str, keep_default_na=False)
                    rows = len(df)
                except Exception:
                    pass
            exports.append({
                "vendor": run.get("vendor", ""),
                "category": run.get("category", ""),
                "path": path,
                "rows": rows,
                "timestamp": run.get("timestamp", ""),
                "stages": run.get("stages", []),
            })
    return exports


def get_today_combined_export(category):
    """Combine all today's exports for a category into one CSV for the row divider."""
    today = datetime.date.today().isoformat()
    export_dir = os.path.join(CLEANED_DIR, today)
    if not os.path.exists(export_dir):
        return pd.DataFrame(), {"error": "No exports for today"}

    frames = []
    files_used = []
    for fname in os.listdir(export_dir):
        if fname.endswith(".csv") and safe_filename(category) in fname:
            fpath = os.path.join(export_dir, fname)
            try:
                df = pd.read_csv(fpath, dtype=str, keep_default_na=False)
                frames.append(df)
                files_used.append(fname)
            except Exception:
                continue

    if not frames:
        return pd.DataFrame(), {"error": f"No exports for {category} today"}

    combined = pd.concat(frames, ignore_index=True)
    before = len(combined)
    combined = combined.drop_duplicates(subset=["Phone Number"], keep="first")

    return combined, {
        "category": category,
        "files": files_used,
        "total_rows": before,
        "cross_vendor_dupes": before - len(combined),
        "final_rows": len(combined),
    }


def _month_tab_name(dt):
    """Generate month tab name like 'Mar 26' from a date."""
    return dt.strftime("%b %y") if hasattr(dt, "strftime") else ""


def _parse_date_value(d):
    """Parse a date value from Excel (datetime) or string formats."""
    if d is None:
        return None
    if isinstance(d, datetime.datetime):
        return d.date()
    if isinstance(d, datetime.date):
        return d
    s = str(d).strip()
    if not s:
        return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y"]:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def get_master_list_categories():
    """Return list of configured master list categories."""
    cats = list(CATEGORY_CONFIGS.keys())
    # Also check for any new .xlsx files in the pipeline_data/master_lists dir
    if os.path.exists(MASTER_LISTS_DIR):
        for f in os.listdir(MASTER_LISTS_DIR):
            if f.endswith(".xlsx"):
                name = f.replace(".xlsx", "")
                if name not in cats:
                    cats.append(name)
    return sorted(cats)


def _get_category_config(category):
    """Get config for a category, with fallback for custom categories."""
    if category in CATEGORY_CONFIGS:
        return CATEGORY_CONFIGS[category]
    # Fallback: look in pipeline_data/master_lists
    return {
        "file": os.path.join(MASTER_LISTS_DIR, f"{safe_filename(category)}.xlsx"),
        "rolling_tab": "30 Days",
        "dedup_days": 30,
        "archive_mode": "move",
    }


def get_master_list(category, sheet=None):
    """
    Get master list data for a category.
    If sheet is None, returns the rolling tab.
    Returns: DataFrame
    """
    import openpyxl
    cfg = _get_category_config(category)
    fpath = cfg["file"]

    if not os.path.exists(fpath):
        return pd.DataFrame(columns=MASTER_LIST_COLS)

    target_sheet = sheet or cfg["rolling_tab"]

    try:
        df = pd.read_excel(fpath, sheet_name=target_sheet, dtype=str, keep_default_na=False)
        # Clean up any extra None columns
        df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
        return df
    except Exception:
        return pd.DataFrame(columns=MASTER_LIST_COLS)


def get_master_list_sheets(category):
    """Return list of sheet names and row counts for a category's workbook."""
    import openpyxl
    cfg = _get_category_config(category)
    fpath = cfg["file"]

    if not os.path.exists(fpath):
        return {"sheets": [], "rolling_tab": cfg["rolling_tab"], "dedup_days": cfg["dedup_days"]}

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            row_count = max(ws.max_row - 1, 0)  # minus header
            sheets.append({"name": name, "rows": row_count})
        wb.close()
        return {
            "sheets": sheets,
            "rolling_tab": cfg["rolling_tab"],
            "dedup_days": cfg["dedup_days"],
            "file": fpath,
        }
    except Exception as e:
        return {"sheets": [], "error": str(e)}


def add_to_master_list(category, cleaned_df):
    """
    Append cleaned data to the rolling tab of the category's Excel workbook.
    Also deduplicates today's data against the existing rolling window.
    Returns stats about the operation.
    """
    import openpyxl
    cfg = _get_category_config(category)
    fpath = cfg["file"]
    rolling_tab = cfg["rolling_tab"]
    dedup_days = cfg["dedup_days"]

    # Ensure cleaned_df has correct columns
    for col in MASTER_LIST_COLS:
        if col not in cleaned_df.columns:
            cleaned_df[col] = ""

    # Read existing rolling tab
    if os.path.exists(fpath):
        try:
            existing = pd.read_excel(fpath, sheet_name=rolling_tab, keep_default_na=False)
            existing = existing.loc[:, ~existing.columns.str.startswith("Unnamed")]
        except Exception:
            existing = pd.DataFrame(columns=MASTER_LIST_COLS)
    else:
        existing = pd.DataFrame(columns=MASTER_LIST_COLS)

    # Get phone numbers already in rolling window for dedup
    existing_phones = set()
    if not existing.empty and "Phone Number" in existing.columns:
        existing_phones = set(existing["Phone Number"].dropna().unique())

    # Dedup today's data against rolling window
    new_data = cleaned_df[MASTER_LIST_COLS].copy()
    before_dedup = len(new_data)

    # Remove rows whose phone is already in rolling tab
    if existing_phones:
        dupe_mask = new_data["Phone Number"].isin(existing_phones)
        removed_as_dupes = int(dupe_mask.sum())
        new_data = new_data[~dupe_mask].copy()
    else:
        removed_as_dupes = 0

    # Also dedupe within today's batch
    before_internal = len(new_data)
    new_data = new_data.drop_duplicates(subset=["Phone Number"], keep="first")
    internal_dupes = before_internal - len(new_data)

    # Convert date strings to datetime for Excel compatibility
    def to_excel_date(val):
        d = _parse_date_value(val)
        if d:
            return datetime.datetime(d.year, d.month, d.day)
        return val

    new_data["Date"] = new_data["Date"].apply(to_excel_date)

    # Append to rolling tab
    if not existing.empty:
        # Convert existing dates too for consistency
        if "Date" in existing.columns:
            existing["Date"] = existing["Date"].apply(to_excel_date)
        combined = pd.concat([existing, new_data], ignore_index=True)
    else:
        combined = new_data.copy()

    # Now handle rotation: remove data older than dedup_days from rolling tab
    today = datetime.date.today()
    cutoff = today - datetime.timedelta(days=dedup_days)
    archive_mode = cfg.get("archive_mode", "move")

    combined["_parsed_date"] = combined["Date"].apply(_parse_date_value)
    old_mask = combined["_parsed_date"].notna() & (combined["_parsed_date"] < cutoff)
    old_data = combined[old_mask].copy()
    keep_data = combined[~old_mask].copy()

    # Group old data by month for archiving
    archived_months = {}
    if not old_data.empty:
        old_data["_month"] = old_data["_parsed_date"].apply(lambda d: _month_tab_name(d) if d else "Unknown")
        for month_name, group in old_data.groupby("_month"):
            if month_name:
                archived_months[month_name] = group.drop(columns=["_parsed_date", "_month"])

    # For "copy" mode (MS/Slice): also copy today's new data to the current month tab
    copy_to_month = {}
    if archive_mode == "copy" and not new_data.empty:
        current_month = _month_tab_name(today)
        if current_month:
            copy_to_month[current_month] = new_data.copy()
            # Merge with any archived data for this month
            if current_month in archived_months:
                copy_to_month[current_month] = pd.concat([archived_months[current_month], new_data], ignore_index=True)
                del archived_months[current_month]

    # Combine archive dicts
    all_archives = {**archived_months, **copy_to_month}

    # Drop helper columns
    keep_data = keep_data.drop(columns=["_parsed_date"], errors="ignore")

    # Write back to Excel workbook
    _write_master_workbook(fpath, rolling_tab, keep_data[MASTER_LIST_COLS], all_archives)

    stats = {
        "category": category,
        "existing_rolling_rows": len(existing),
        "new_rows_before_dedup": before_dedup,
        "removed_as_dupes": removed_as_dupes,
        "internal_dupes": internal_dupes,
        "added_rows": len(new_data),
        "rolling_tab_rows": len(keep_data),
        "archived_months": {k: len(v) for k, v in archived_months.items()},
        "copied_to_month": {k: len(v) for k, v in copy_to_month.items()},
        "archive_mode": archive_mode,
        "dedup_days": dedup_days,
    }

    return stats


def dedup_master_list(category, days=None):
    """
    Run deduplication on the rolling tab. This is called after adding new data.
    Reads the rolling tab, dedupes today's entries against the window, and saves back.
    Also rotates old data to monthly archive tabs.
    Returns: (today_kept_df, removed_df, stats)
    """
    import openpyxl
    cfg = _get_category_config(category)
    fpath = cfg["file"]
    rolling_tab = cfg["rolling_tab"]
    dedup_days = days or cfg["dedup_days"]

    if not os.path.exists(fpath):
        return pd.DataFrame(), pd.DataFrame(), {"error": "Master list file not found"}

    # Read rolling tab
    try:
        rolling = pd.read_excel(fpath, sheet_name=rolling_tab, keep_default_na=False)
        rolling = rolling.loc[:, ~rolling.columns.str.startswith("Unnamed")]
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame(), {"error": f"Cannot read rolling tab: {str(e)}"}

    if rolling.empty:
        return pd.DataFrame(), pd.DataFrame(), {"error": "Rolling tab is empty"}

    today = datetime.date.today()
    cutoff = today - datetime.timedelta(days=dedup_days)

    # Parse dates
    rolling["_parsed_date"] = rolling["Date"].apply(_parse_date_value)

    # Split: today's rows vs historical (within window) vs old (outside window)
    has_date = rolling["_parsed_date"].notna()
    today_mask = has_date & (rolling["_parsed_date"] == today)
    recent_mask = has_date & (rolling["_parsed_date"] >= cutoff) & (rolling["_parsed_date"] < today)
    old_mask = has_date & (rolling["_parsed_date"] < cutoff)

    # Get phone numbers from recent history (not today)
    recent_phones = set()
    if recent_mask.any():
        recent_phones = set(rolling.loc[recent_mask, "Phone Number"].dropna().unique())

    # Dedup today's rows against recent
    today_rows = rolling[today_mask].copy()
    removed = pd.DataFrame()
    internal_dupes = 0

    if not today_rows.empty:
        if recent_phones:
            dupe_mask = today_rows["Phone Number"].isin(recent_phones)
            removed = today_rows[dupe_mask].copy()
            kept_today = today_rows[~dupe_mask].copy()
        else:
            kept_today = today_rows.copy()

        # Internal dedup within today
        before_internal = len(kept_today)
        kept_today = kept_today.drop_duplicates(subset=["Phone Number"], keep="first")
        internal_dupes = before_internal - len(kept_today)
    else:
        kept_today = pd.DataFrame()

    # Archive old data to monthly tabs
    archive_mode = cfg.get("archive_mode", "move")
    old_data = rolling[old_mask].copy()
    archived_months = {}
    if not old_data.empty:
        old_data["_month"] = old_data["_parsed_date"].apply(lambda d: _month_tab_name(d) if d else "")
        for month_name, group in old_data.groupby("_month"):
            if month_name:
                # For "move" mode (SFG): archive old data to month tab (it leaves the rolling tab)
                # For "copy" mode (MS): old data was already copied daily, so just remove from rolling
                if archive_mode == "move":
                    archived_months[month_name] = group.drop(columns=["_parsed_date", "_month"], errors="ignore")
                # For "copy" mode: don't re-archive (it's already in the month tab from daily copy)

    # Rebuild rolling tab: recent + deduped today (no old data)
    parts = []
    if recent_mask.any():
        parts.append(rolling[recent_mask].drop(columns=["_parsed_date"], errors="ignore"))
    if not kept_today.empty:
        parts.append(kept_today.drop(columns=["_parsed_date"], errors="ignore"))
    # Also include any rows without valid dates
    no_date_mask = ~has_date
    if no_date_mask.any():
        parts.append(rolling[no_date_mask].drop(columns=["_parsed_date"], errors="ignore"))

    new_rolling = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=MASTER_LIST_COLS)

    # Clean up columns
    for col in ["_parsed_date", "_month"]:
        if col in new_rolling.columns:
            new_rolling.drop(columns=[col], inplace=True)
        if col in removed.columns:
            removed.drop(columns=[col], inplace=True)

    # Write back
    _write_master_workbook(fpath, rolling_tab, new_rolling[MASTER_LIST_COLS] if not new_rolling.empty else pd.DataFrame(columns=MASTER_LIST_COLS), archived_months)

    stats = {
        "category": category,
        "rolling_tab": rolling_tab,
        "archive_mode": archive_mode,
        "total_in_rolling": len(rolling),
        "recent_phones_in_window": len(recent_phones),
        "today_rows": len(today_rows),
        "removed_as_dupes": len(removed),
        "internal_dupes_removed": internal_dupes,
        "kept_today": len(kept_today) if not kept_today.empty else 0,
        "old_rows_removed": len(old_data),
        "archived_to_months": {k: len(v) for k, v in archived_months.items()},
        "new_rolling_rows": len(new_rolling),
        "dedup_window_days": dedup_days,
    }

    # The "today kept" data is what gets exported for HeyMarket/GoHighLevel
    today_export = kept_today.drop(columns=["_parsed_date", "_month"], errors="ignore") if not kept_today.empty else pd.DataFrame(columns=MASTER_LIST_COLS)

    return today_export, removed.drop(columns=["_parsed_date", "_month"], errors="ignore") if not removed.empty else pd.DataFrame(), stats


def _write_master_workbook(fpath, rolling_tab_name, rolling_df, archived_months):
    """
    Write the master workbook:
    - Replace the rolling tab with new data
    - Append archived data to the correct monthly tabs (create if needed)
    Preserves all existing monthly archive tabs.
    """
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    if os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = rolling_tab_name

    # Update rolling tab
    if rolling_tab_name in wb.sheetnames:
        # Clear and rewrite
        ws = wb[rolling_tab_name]
        # Delete all rows except keep the sheet
        ws.delete_rows(1, ws.max_row + 1)
    else:
        ws = wb.create_sheet(rolling_tab_name, 0)

    # Write header
    for col_idx, col_name in enumerate(MASTER_LIST_COLS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    if not rolling_df.empty:
        for row_idx, (_, row) in enumerate(rolling_df.iterrows(), 2):
            for col_idx, col_name in enumerate(MASTER_LIST_COLS, 1):
                val = row.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Ensure rolling tab is first
    if wb.sheetnames.index(rolling_tab_name) != 0:
        wb.move_sheet(rolling_tab_name, offset=-wb.sheetnames.index(rolling_tab_name))

    # Append to monthly archive tabs
    for month_name, month_df in archived_months.items():
        if not month_name or month_df.empty:
            continue

        if month_name in wb.sheetnames:
            # Append to existing tab
            ws_month = wb[month_name]
            start_row = ws_month.max_row + 1
        else:
            # Create new tab
            ws_month = wb.create_sheet(month_name)
            # Write header
            for col_idx, col_name in enumerate(MASTER_LIST_COLS, 1):
                ws_month.cell(row=1, column=col_idx, value=col_name)
            start_row = 2

        # Write month data
        for row_idx, (_, row) in enumerate(month_df.iterrows(), start_row):
            for col_idx, col_name in enumerate(MASTER_LIST_COLS, 1):
                val = row.get(col_name, "")
                ws_month.cell(row=row_idx, column=col_idx, value=val)

    wb.save(fpath)
    wb.close()


# =============================================================================
# ROW DIVIDER
# =============================================================================

DEFAULT_ACCOUNTS = [
    {"name": "Griffin Insurance", "amount": 275},
    {"name": "GKM Insurance", "amount": 275},
    {"name": "EA Insurance", "amount": 175},
    {"name": "Spencer Financial", "amount": 175},
    {"name": "Dilbert Financial", "amount": 150},
    {"name": "SWAT Financial", "amount": 125},
    {"name": "Else Agency", "amount": 125},
    {"name": "Manzano Agency", "amount": 100},
    {"name": "King Agency", "amount": 75},
    {"name": "Montgomery Agency", "amount": 50},
]


def get_pipeline_config():
    """Load pipeline configuration (accounts, categories, etc.)."""
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r") as f:
                return json.load(f)
        except Exception:
            pass

    # Default config
    return {
        "accounts": DEFAULT_ACCOUNTS,
        "categories": [
            {"key": "SFG", "label": "SFG (Indeed)", "dedup_days": 30, "rolling_tab": "30 Days",
             "file": "Indeed_SFG_MasterList.xlsx", "archive_mode": "move"},
            {"key": "MS", "label": "MS / Slice (Indeed)", "dedup_days": 60, "rolling_tab": "60 Days",
             "file": "Slice_Indeed_MasterList.xlsx", "archive_mode": "copy"},
        ],
        "list_template": "{date} - {account} - Indeed",
        "default_company": "FBSPL",
        "default_stage": "Initial",
        "assignment_style": "smart_random",
        "row_mode": "auto_scale",
    }


def save_pipeline_config(config):
    """Save pipeline configuration."""
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=2)
    return {"ok": True}


def _normalize_counts_to_total(weights, total_rows):
    total_weight = sum(v for v in weights.values() if v > 0)
    if total_weight <= 0:
        raise ValueError("Amounts must sum to a positive number.")
    scaled = {k: (v / total_weight) * total_rows for k, v in weights.items()}
    rounded = {k: int(round(v)) for k, v in scaled.items()}
    diff = total_rows - sum(rounded.values())
    keys = list(rounded.keys())
    i = 0
    while diff != 0:
        k = keys[i % len(keys)]
        rounded[k] += 1 if diff > 0 else -1
        diff += -1 if diff > 0 else 1
        i += 1
    return rounded


def _build_account_sequence(counts, style):
    accounts = [a for a, c in counts.items() if c > 0]
    if style == "top_to_bottom":
        seq = []
        for a in accounts:
            seq.extend([a] * counts[a])
        return seq
    # Smart random mix
    remaining = counts.copy()
    seq, last = [], None
    while sum(remaining.values()) > 0:
        candidates = [a for a in accounts if remaining[a] > 0]
        candidates.sort(key=lambda a: remaining[a], reverse=True)
        bucket = candidates[:min(4, len(candidates))]
        if last in bucket and len(bucket) > 1:
            bucket.remove(last)
        pick = random.choice(bucket)
        seq.append(pick)
        remaining[pick] -= 1
        last = pick
    return seq


def divide_rows(csv_bytes, accounts, date_val="", list_template="", mode="auto_scale", style="smart_random"):
    """
    Divide CSV rows among accounts.
    accounts: list of {"name": str, "amount": int, "list_name": str (optional)}
    mode: "auto_scale" | "exact_with_unused" | "exact_match"
    style: "smart_random" | "top_to_bottom"
    Returns: (zip_bytes, stats)
    """
    if not date_val:
        today = datetime.date.today()
        date_val = f"{today.month}/{today.day}/{today.year}"

    if not list_template:
        list_template = f"{date_val} - {{account}} - Indeed"

    # Read CSV
    df = read_csv_robust(csv_bytes)
    total_rows = len(df)

    # Build allocation dict
    alloc = {}
    list_map = {}
    for a in accounts:
        name = a["name"].strip()
        if not name:
            continue
        amt = int(a.get("amount", 0))
        alloc[name] = amt
        custom_list = a.get("list_name", "").strip()
        list_map[name] = custom_list if custom_list else list_template.replace("{account}", name)

    if not alloc or sum(alloc.values()) == 0:
        return None, {"error": "No valid accounts with positive amounts"}

    requested_total = sum(alloc.values())

    if mode == "auto_scale":
        final_counts = _normalize_counts_to_total(alloc, total_rows)
        used_rows = total_rows
        unused_rows = 0
    elif mode == "exact_with_unused":
        if requested_total > total_rows:
            return None, {"error": f"Requested {requested_total} rows but only {total_rows} exist"}
        final_counts = alloc.copy()
        used_rows = requested_total
        unused_rows = total_rows - requested_total
    else:  # exact_match
        if requested_total != total_rows:
            return None, {"error": f"Exact match required. Requested {requested_total}, file has {total_rows}"}
        final_counts = alloc.copy()
        used_rows = total_rows
        unused_rows = 0

    # Build assignment sequence
    sequence = _build_account_sequence(final_counts, style)

    used_df = df.iloc[:used_rows].copy()
    unused_df = df.iloc[used_rows:].copy() if unused_rows else df.iloc[0:0].copy()

    if "List" not in used_df.columns:
        used_df["List"] = ""
    if "Date" not in used_df.columns:
        used_df["Date"] = ""

    used_df["_acct"] = sequence[:used_rows]
    used_df["List"] = used_df["_acct"].apply(lambda a: list_map.get(a, a))
    used_df["Date"] = date_val
    used_df.drop(columns="_acct", inplace=True)

    # Build ZIP
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("MASTER_used.csv", used_df.to_csv(index=False).encode("utf-8"))
        for acct in final_counts:
            acct_df = used_df[used_df["List"] == list_map[acct]]
            z.writestr(f"{safe_filename(acct)}.csv", acct_df.to_csv(index=False).encode("utf-8"))
        if unused_rows:
            z.writestr("UNUSED.csv", unused_df.to_csv(index=False).encode("utf-8"))

    stats = {
        "total_rows": total_rows,
        "used_rows": used_rows,
        "unused_rows": unused_rows,
        "accounts": {k: v for k, v in final_counts.items()},
        "mode": mode,
    }

    return buf.getvalue(), stats


# =============================================================================
# EXTRACT CSVs FROM VENDOR ZIP
# =============================================================================

def extract_csvs_from_zip(zip_bytes):
    """
    Extract CSV files from a vendor upload ZIP.
    Returns list of (filename, csv_bytes) tuples.
    """
    csvs = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as z:
        for name in z.namelist():
            lower = name.lower()
            if lower.endswith((".csv", ".xlsx", ".xls")) and not name.startswith("__MACOSX"):
                csvs.append((os.path.basename(name), z.read(name)))
    return csvs


# =============================================================================
# FULL PIPELINE: Extract → Clean → Add to Master → Dedup → Divide
# =============================================================================

def run_full_pipeline(zip_bytes_or_csv_list, category, cleaner_type="indeed",
                      company="", list_val="", date_val="", stage="Initial",
                      do_dedup=True, dedup_days=30,
                      do_divide=False, accounts=None, list_template="",
                      divide_mode="auto_scale", divide_style="smart_random"):
    """
    Run the full pipeline on uploaded data.
    Returns a results dict with stats from each stage.
    """
    results = {"stages": []}

    # Stage 1: Extract CSVs if ZIP
    if isinstance(zip_bytes_or_csv_list, bytes):
        csv_list = extract_csvs_from_zip(zip_bytes_or_csv_list)
        results["stages"].append({
            "name": "extract",
            "files_found": len(csv_list),
            "filenames": [f[0] for f in csv_list],
        })
    else:
        csv_list = zip_bytes_or_csv_list

    if not csv_list:
        results["error"] = "No CSV files found"
        return results

    # Stage 2: Clean
    if cleaner_type == "linkedin":
        cleaned_df, clean_stats = clean_linkedin(csv_list, company=company, list_val=list_val,
                                                  date_val=date_val, stage=stage)
    else:
        cleaned_df, clean_stats = clean_indeed(csv_list, company=company, list_val=list_val,
                                                date_val=date_val, stage=stage)

    results["stages"].append({"name": "clean", "type": cleaner_type, **clean_stats})

    if cleaned_df is None or cleaned_df.empty:
        results["error"] = "Cleaning produced no results"
        return results

    # Save cleaned output
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    clean_dir = os.path.join(CLEANED_DIR, today_str)
    os.makedirs(clean_dir, exist_ok=True)
    clean_path = os.path.join(clean_dir, f"{safe_filename(category)}_{cleaner_type}_cleaned.csv")
    cleaned_df.to_csv(clean_path, index=False)
    results["cleaned_path"] = clean_path
    results["cleaned_csv"] = cleaned_df.to_csv(index=False)

    # Stage 3: Add to master list
    add_stats = add_to_master_list(category, cleaned_df)
    results["stages"].append({"name": "master_list_add", **add_stats})

    # Stage 4: Dedup
    if do_dedup:
        deduped_df, removed_df, dedup_stats = dedup_master_list(category, days=dedup_days)
        results["stages"].append({"name": "dedup", **dedup_stats})

        # The deduped "today" rows are the ones we'd send to row divider
        # Extract just today's kept rows for division
        today = datetime.date.today()
        today_str_short = f"{today.month}/{today.day}/{today.year}"

        today_rows = deduped_df[deduped_df.get("Date", "") == today_str_short] if "Date" in deduped_df.columns else pd.DataFrame()

        if today_rows.empty:
            # Try other date format
            today_str_alt = today.strftime("%Y-%m-%d")
            today_rows = deduped_df[deduped_df.get("Date", "") == today_str_alt] if "Date" in deduped_df.columns else pd.DataFrame()

        results["deduped_today_rows"] = len(today_rows)
        results["deduped_today_csv"] = today_rows.to_csv(index=False) if not today_rows.empty else ""
        results["removed_csv"] = removed_df.to_csv(index=False) if not removed_df.empty else ""
    else:
        today_rows = cleaned_df
        results["deduped_today_rows"] = len(today_rows)
        results["deduped_today_csv"] = today_rows.to_csv(index=False)

    # Stage 5: Row Divide
    if do_divide and accounts and not today_rows.empty:
        csv_bytes = today_rows.to_csv(index=False).encode("utf-8")
        zip_bytes, divide_stats = divide_rows(
            csv_bytes, accounts, date_val=date_val,
            list_template=list_template, mode=divide_mode, style=divide_style
        )
        results["stages"].append({"name": "divide", **divide_stats})
        if zip_bytes:
            # Save divided output
            divide_path = os.path.join(DIVIDED_DIR, f"{today_str}_{safe_filename(category)}_divided.zip")
            with open(divide_path, "wb") as f:
                f.write(zip_bytes)
            results["divided_zip_path"] = divide_path

    results["ok"] = True
    return results
